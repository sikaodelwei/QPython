#-*-coding:utf8-*-
from openpyxl import load_workbook
import time
list=[]
listC=[]

wb = load_workbook('../Download/WeiXin/贵阳南国花锦店_书店陈列结果(1).xlsx')
table = wb['数据']

def search(isbn):
  for r in range(1,table.max_row): 
     table_isbn = table.cell(r,3).value     
     if isbn == table_isbn:
        r=r
        table_name = table.cell(r,2).value
        table_isbn = table.cell(r,3).value
        table_loc  = table.cell(r,8).value
        table_desk = table.cell(r,9).value
        table_stock =table.cell(r,10).value
   
        list.append(table_desk+' '+table_name+table_loc)
        list.sort()
     if isbn != table_isbn :
        temp_isbn=isbn
  listC.append(temp_isbn) 
         

'''
办公室扫码枪扫码到微信文件传输手，
粘贴书号到BookNo
'''
import os
f=open('BookNo.txt')
list_txt=f.readlines()
for isbn in list_txt:
    search(isbn)
'''    
结果输出
'''
print("\n\n")
for l in list:
    print(l)
print("\n\n剩下孤儿书在C区")
